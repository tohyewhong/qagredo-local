#!/usr/bin/env python3
"""
Multi-format input normalizer for QAGRedo.

Converts: pdf | txt | xlsx | json | jsonl
Outputs: JSONL where each line is a normalized document record that QAGRedo can ingest.

Canonical record schema (per line):
  - id
  - title
  - content  (preferred by QAGRedo)
  - text     (same as content; included for clarity)
  - source
  - type
  - metadata (optional)

Examples:
  /home/tyewhong/qagredo/.venv/bin/python scripts/conversion/convert_to_qagredo_jsonl.py --input data/dev-data.jsonl --output data/normalized.jsonl
  /home/tyewhong/qagredo/.venv/bin/python scripts/conversion/convert_to_qagredo_jsonl.py --input data/my.pdf --output data/my.jsonl
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional


SUPPORTED_INPUT_TYPES = ("pdf", "txt", "xlsx", "json", "jsonl")


def _repo_root() -> Path:
    """
    Find the QAGRedo repo root even when this script lives under /scripts/.

    We locate the first parent directory that contains:
      - utils/parse.py
      - run_qa_pipeline.py
    """
    here = Path(__file__).resolve()
    # Search upward a few levels (this script is typically at scripts/conversion/*.py)
    for parent in [here.parent, *list(here.parents)[:8]]:
        if (parent / "utils" / "parse.py").is_file() and (parent / "run_qa_pipeline.py").is_file():
            return parent
    # Fallback: best-effort (should still work when run from repo root cwd)
    return here.parent.parent


_REPO_ROOT = _repo_root()
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))


from utils.parse import load_article_json  # noqa: E402


def _resolve_input_path(file_path: str) -> Path:
    """
    Resolve a user-provided file path in a layman-friendly way.

    Tries:
    - absolute path as-is
    - relative path as-is
    - under ./data/
    """
    path = Path(file_path)
    if path.is_absolute() and path.exists():
        return path

    candidates: List[Path] = []
    candidates.append((_REPO_ROOT / path).resolve())
    candidates.append((_REPO_ROOT / "data" / path).resolve())

    for candidate in candidates:
        if candidate.exists() and candidate.is_file():
            return candidate

    tried = "\n".join(f"  - {c}" for c in candidates)
    raise FileNotFoundError(f"File not found: {file_path}\nTried locations:\n{tried}")


def _safe_slug(value: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", value).strip("_")
    return slug or "doc"


# ---------------------------------------------------------------------------
#  JSON repair — tolerate common hand-editing mistakes
# ---------------------------------------------------------------------------

def _try_repair_json(raw_text: str) -> Any:
    """
    Parse JSON text, falling back to lightweight repair for common errors.

    Repairs handled (in order):
      1. Unterminated strings (odd unescaped-quote count on a line)
      2. Missing commas between adjacent object properties / array elements
      3. Trailing commas before ``]`` or ``}``
    """
    try:
        return json.loads(raw_text)
    except json.JSONDecodeError:
        pass

    # --- Repair pass 1 — close unterminated strings ---
    # JSON does not allow multi-line strings, so every line must have an
    # even number of unescaped double-quotes.  An odd count means a
    # string was opened but never closed on that line.
    fixed_lines: List[str] = []
    for line in raw_text.split("\n"):
        stripped = line.rstrip()
        n_quotes = 0
        i = 0
        while i < len(stripped):
            if stripped[i] == "\\" and i + 1 < len(stripped):
                i += 2  # skip escaped character
                continue
            if stripped[i] == '"':
                n_quotes += 1
            i += 1
        if n_quotes % 2 == 1:
            # If line ends with a trailing comma, close the string *before*
            # the comma so the comma stays a JSON delimiter.
            # e.g.  "title": "abc,  →  "title": "abc",
            trail_comma = re.search(r',\s*$', stripped)
            if trail_comma:
                stripped = stripped[:trail_comma.start()] + '",'
            else:
                stripped += '"'
        fixed_lines.append(stripped)
    repaired = "\n".join(fixed_lines)

    # --- Repair pass 2 — insert missing commas ---
    # A line that ends with a value token (string / number / bool / null / ] / })
    # followed (after whitespace + newline) by a line starting with '"' (a new key)
    # without an intervening comma.
    repaired = re.sub(
        r'("(?:[^"\\]|\\.)*"|true|false|null|\d+(?:\.\d+)?|\]|\})\s*\n(\s*")',
        r"\1,\n\2",
        repaired,
    )

    # --- Repair pass 3 — strip trailing commas before ] or } ---
    repaired = re.sub(r",\s*([\]\}])", r"\1", repaired)

    try:
        return json.loads(repaired)
    except json.JSONDecodeError as exc:
        raise json.JSONDecodeError(
            f"Could not parse or auto-repair JSON: {exc.msg}", exc.doc, exc.pos
        ) from exc


# ---------------------------------------------------------------------------
#  Fallback JSON object extraction from broken text
# ---------------------------------------------------------------------------

def _regex_extract_press_doc(block: str) -> Optional[Dict[str, Any]]:
    """
    Last-resort extraction: pull press-style document fields out of a
    ``{…}`` text block using regex when JSON parsing has failed completely.

    Extracts ``country``, ``title``, ``summary``, **English-only** ``"article"``
    values, and ``source_date`` entries.  Articles inside ``"native"`` blocks
    are skipped.
    """
    doc: Dict[str, Any] = {}

    # -- scalar fields --
    for key in ("country", "title", "summary"):
        m = re.search(rf'"{key}"\s*:\s*"((?:[^"\\]|\\.)*)"', block)
        if m:
            doc[key] = m.group(1)

    # -- English-only article texts (skip articles inside "native" blocks) --
    # Find positions of all "native" keys so we can exclude articles near them.
    native_positions = [m.start() for m in re.finditer(r'"native"\s*:', block)]
    all_article_matches = list(re.finditer(r'"article"\s*:\s*"((?:[^"\\]|\\.)*)"', block))

    articles: List[str] = []
    for m in all_article_matches:
        art_text = m.group(1).strip()
        if not art_text:
            continue
        # Skip if a "native" key appears within 200 chars before this article
        # (i.e. this article is inside a native block)
        is_native = any(0 < m.start() - np < 200 for np in native_positions)
        if not is_native:
            articles.append(art_text)

    if articles:
        source_items: List[Dict[str, Any]] = []
        for art_text in articles:
            source_items.append({"english": {"article": art_text}})
        doc["source"] = source_items

    # -- source_date entries --
    # Matches {…"source":"X"…"day":N…"month":"X"…"year":N…}
    sd_blocks = re.finditer(
        r'\{\s*"source"\s*:\s*"((?:[^"\\]|\\.)*)"\s*,\s*'
        r'"day"\s*:\s*(\d+)\s*,\s*'
        r'"month"\s*:\s*"((?:[^"\\]|\\.)*)"\s*,\s*'
        r'"year"\s*:\s*(\d+)',
        block,
    )
    source_dates: List[Dict[str, Any]] = []
    for m in sd_blocks:
        source_dates.append({
            "source": m.group(1),
            "day": int(m.group(2)),
            "month": m.group(3),
            "year": int(m.group(4)),
        })
    if source_dates:
        doc["source_date"] = source_dates

    # Only return if we extracted something useful
    if not doc.get("source") and not doc.get("summary"):
        return None
    return doc


def _extract_json_objects_from_text(raw_text: str) -> List[Dict[str, Any]]:
    """
    Fallback extractor: scan *raw_text* for top-level ``{…}`` blocks
    and attempt to parse each one individually (with repair).

    If a block still cannot be parsed after repair, a regex-based
    extraction is tried as a last resort (for press-style documents).

    This is useful when the overall JSON structure (e.g. the wrapping
    ``[…]`` array) is broken but individual document objects are still
    salvageable.
    """
    docs: List[Dict[str, Any]] = []
    i = 0
    length = len(raw_text)

    while i < length:
        if raw_text[i] != '{':
            i += 1
            continue

        # Track brace depth to locate the matching ``}``
        depth = 0
        start = i
        in_string = False
        escape_next = False
        j = i

        while j < length:
            c = raw_text[j]
            if escape_next:
                escape_next = False
                j += 1
                continue
            if in_string:
                if c == '\\':
                    escape_next = True
                elif c == '"':
                    in_string = False
            else:
                if c == '"':
                    in_string = True
                elif c == '{':
                    depth += 1
                elif c == '}':
                    depth -= 1
                    if depth == 0:
                        block = raw_text[start:j + 1]
                        parsed = False
                        try:
                            obj = _try_repair_json(block)
                            if isinstance(obj, dict):
                                docs.append(obj)
                                parsed = True
                        except Exception:
                            pass
                        # Last resort: regex extraction for press-style docs
                        if not parsed:
                            regex_doc = _regex_extract_press_doc(block)
                            if regex_doc:
                                docs.append(regex_doc)
                        i = j + 1
                        break
            j += 1
        else:
            # Reached end without finding matching brace
            break

        if depth != 0:
            i = start + 1  # skip this ``{`` and try the next one

    return docs


# ---------------------------------------------------------------------------
#  Press / news-article schema helpers
# ---------------------------------------------------------------------------

def _get_source_list(doc: Dict[str, Any]) -> Optional[list]:
    """Return the source list from *doc*, tolerating both ``source`` and ``sources``."""
    for key in ("source", "sources"):
        val = doc.get(key)
        if isinstance(val, list):
            return val
    return None


def _get_source_dates(doc: Dict[str, Any]) -> list:
    """Return the source-date list from *doc*, tolerating both ``source_date`` and ``source_dates``."""
    for key in ("source_date", "source_dates"):
        val = doc.get(key)
        if isinstance(val, list):
            return val
    return []


def _is_press_style_doc(doc: Dict[str, Any]) -> bool:
    """
    Detect whether *doc* follows the press / news-article schema.

    Indicators:
      * ``source`` or ``sources`` is a **list** (language-variant objects, not a file path)
      * at least one of ``country``, ``summary``, ``source_date``/``source_dates`` is present
    """
    if _get_source_list(doc) is None:
        return False
    return any(k in doc for k in ("country", "summary", "source_date", "source_dates"))


def _extract_source_articles(src_item: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Extract **English-only** article text from a single item in the ``source`` array.

    All ``"native"`` content is **skipped** — regardless of whether it is
    ``null``, ``{}`` (empty), or contains actual article text.  Only the
    ``"english"`` article (or a flat/unwrapped ``"article"``) is included.

    Tolerates several structures that arise from hand-editing:

    1. **Standard** — ``english`` wrapper present::

           {"english": {"article": "…"}, "native": …}   → only english used

    2. **Flat / unwrapped** — no language key at all::

           {"article": "…", "title": "…"}               → used as-is

    Returns a list of dicts each with keys:
    ``language``, ``article``, ``title``, ``source_date`` (any may be empty/None).
    """
    results: List[Dict[str, Any]] = []
    if not isinstance(src_item, dict):
        return results

    found_via_lang_key = False

    # ---- Only extract "english" articles; skip ALL "native" content ----
    # "native" is excluded regardless of whether it is null, {}, or has
    # actual article text.  Only English-language articles go into content.
    lang_data = src_item.get("english")
    if isinstance(lang_data, dict) and lang_data:
        article_text = lang_data.get("article", "")
        if isinstance(article_text, str) and article_text.strip():
            results.append({
                "language": "english",
                "article": article_text.strip(),
                "title": str(lang_data.get("title", "") or ""),
                "source_date": lang_data.get("source_date"),
            })
            found_via_lang_key = True

    # Flat / unwrapped structure — ``article`` lives directly on the source item
    # (no "english" wrapper at all).  "native" siblings are ignored.
    if not found_via_lang_key:
        direct_article = src_item.get("article", "")
        if isinstance(direct_article, str) and direct_article.strip():
            results.append({
                "language": "unknown",
                "article": direct_article.strip(),
                "title": str(src_item.get("title", "") or ""),
                "source_date": src_item.get("source_date"),
            })

    return results


def _flatten_press_doc(doc: Dict[str, Any], *, doc_index: int = 1) -> Dict[str, Any]:
    """
    Transform a press / news-style nested document into a **flat** dict
    with ``content`` at the top level so the normal QAGRedo text
    extraction works correctly.

    Handles **multiple sources per title**: each source item in the
    ``source`` array may contribute one or more articles (English,
    native, or both).  All article texts are merged into a single
    ``content`` string.

    Content is built **only** from ``article`` fields inside the
    ``source`` array.  The ``summary`` field is preserved in metadata
    but is *not* included in ``content``.

    Metadata preserves: country, summary, source_date, source labels,
    languages, per-source titles, and source count.
    """
    content_parts: List[str] = []

    # -- extract articles from all source items --
    source_list = _get_source_list(doc) or []
    all_articles: List[Dict[str, Any]] = []

    for src_item in source_list:
        all_articles.extend(_extract_source_articles(src_item))

    num_articles = len(all_articles)

    # -- assemble content (from articles only, merged into one) --
    for art_info in all_articles:
        if art_info.get("article"):
            content_parts.append(art_info["article"])

    content = "\n\n".join(content_parts).strip()

    # -- fallback: deep extraction when structured parsing yields nothing --
    if not content and source_list:
        deep_texts = _deep_extract_articles(source_list)
        if deep_texts:
            content = "\n\n".join(deep_texts).strip()

    # -- rich metadata --
    meta: Dict[str, Any] = {}
    if doc.get("country"):
        meta["country"] = doc["country"]

    # Preserve summary in metadata (not in content)
    summary = doc.get("summary", "")
    if isinstance(summary, str) and summary.strip():
        meta["summary"] = summary.strip()

    # Top-level source_date / source_dates array
    source_dates = _get_source_dates(doc)
    if source_dates:
        date_strs: List[str] = []
        source_labels: List[str] = []
        for sd in source_dates:
            if not isinstance(sd, dict):
                continue
            day = sd.get("day", "")
            month = sd.get("month", "")
            year = sd.get("year", "")
            date_strs.append(f"{day}-{month}-{year}")
            if sd.get("source"):
                source_labels.append(str(sd["source"]))
        if date_strs:
            meta["source_date"] = date_strs
        if source_labels:
            meta["source_label"] = source_labels

    # Per-source metadata
    if all_articles:
        languages_found = list(dict.fromkeys(
            a["language"] for a in all_articles if a.get("language")
        ))
        if languages_found:
            meta["languages"] = languages_found
        meta["num_sources"] = num_articles

        source_titles = [a["title"] for a in all_articles if a.get("title")]
        if source_titles:
            meta["source_titles"] = source_titles

        # Dates from within individual source items
        per_source_dates: List[str] = []
        for art_info in all_articles:
            sd = art_info.get("source_date")
            if isinstance(sd, dict):
                day = sd.get("day", "")
                month = sd.get("month", "")
                year = sd.get("year", "")
                per_source_dates.append(f"{day}-{month}-{year}")
        if per_source_dates:
            meta["per_source_dates"] = per_source_dates

    flat: Dict[str, Any] = {
        "title": doc.get("title", f"press_doc_{doc_index}"),
        "content": content,
    }
    if meta:
        flat["metadata"] = meta
    return flat


# ---------------------------------------------------------------------------
#  Deep (recursive) text extraction for arbitrary nested JSON
# ---------------------------------------------------------------------------

def _deep_extract_articles(obj: Any, *, _depth: int = 0) -> List[str]:
    """
    Recursively walk *obj* looking for string values under keys that
    typically hold article / document text (``article``, ``content``,
    ``text``, ``body``, ``passage``).

    **Skips the ``"native"`` key entirely** — only English / non-native
    content is extracted.

    Returns a list of non-empty text fragments found.
    """
    _ARTICLE_KEYS = {"article", "content", "text", "body", "passage"}
    if _depth > 6:
        return []
    results: List[str] = []
    if isinstance(obj, dict):
        for key, val in obj.items():
            # Skip all "native" content (null, empty, or with text)
            if key == "native":
                continue
            if key in _ARTICLE_KEYS and isinstance(val, str) and val.strip():
                results.append(val.strip())
            elif isinstance(val, (dict, list)):
                results.extend(_deep_extract_articles(val, _depth=_depth + 1))
    elif isinstance(obj, list):
        for item in obj:
            results.extend(_deep_extract_articles(item, _depth=_depth + 1))
    return results


def _extract_text_field(document: Dict[str, Any]) -> str:
    """
    Extract text content from a document dict, mirroring QAGRedo’s tolerance.
    """
    text_fields = ["content", "text", "body", "document", "article", "passage", "summary"]
    for field in text_fields:
        if field in document and document[field]:
            val = document[field]
            if isinstance(val, list):
                return " ".join(str(item) for item in val if item is not None)
            if isinstance(val, str) and val.strip():
                return val

    # Deep extraction: search nested dicts/lists for article text
    deep_texts = _deep_extract_articles(document)
    if deep_texts:
        return "\n\n".join(deep_texts)

    # Last resort: concatenate scalar non-metadata fields
    metadata_fields = {"id", "title", "source", "type", "metadata"}
    parts: List[str] = []
    for k, v in document.items():
        if k in metadata_fields or v is None:
            continue
        if isinstance(v, (dict, list)):
            continue
        s = str(v).strip()
        if s:
            parts.append(s)
    return " ".join(parts).strip()


def _canonicalize_record(
    raw: Dict[str, Any],
    *,
    fallback_id: str,
    fallback_title: str,
    source: str,
    input_type: str,
    extra_metadata: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    content = _extract_text_field(raw).strip()
    if not content:
        raise ValueError(f"Empty text after extraction for source={source!r}")

    metadata: Dict[str, Any] = {"input_type": input_type}
    if extra_metadata:
        metadata.update(extra_metadata)
    # Preserve any per-record source for debugging, but standardize `source`
    # to the actual input path to be predictable for layman users.
    raw_source = raw.get("source")
    if raw_source and not isinstance(raw_source, list) and str(raw_source) != str(source):
        metadata.setdefault("original_source", str(raw_source))
    if isinstance(raw.get("metadata"), dict):
        # Merge in any existing metadata without losing our input_type marker.
        merged = dict(raw["metadata"])
        merged.update(metadata)
        metadata = merged

    rec = {
        "id": str(raw.get("id") or fallback_id),
        "title": str(raw.get("title") or fallback_title),
        "content": content,
        "text": content,  # explicit alias; QAGRedo prefers `content` but accepts `text`
        "source": str(source),
        "type": str(raw.get("type") or "text_document"),
    }
    if metadata:
        rec["metadata"] = metadata
    return rec


def _extract_txt(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def _extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except Exception as e:  # pragma: no cover
        raise RuntimeError(
            "PDF support requires the 'pypdf' package. "
            "Install it (or run `pip install -r requirements.txt`)."
        ) from e

    reader = PdfReader(str(path))
    parts: List[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        text = text.strip()
        if text:
            parts.append(text)
    return "\n\n".join(parts).strip()


def _extract_xlsx(path: Path) -> tuple[str, Dict[str, Any]]:
    try:
        import openpyxl
    except Exception as e:  # pragma: no cover
        raise RuntimeError(
            "XLSX support requires the 'openpyxl' package. "
            "Install it (or run `pip install -r requirements.txt`)."
        ) from e

    wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)

    chunks: List[str] = []
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        chunks.append(f"Sheet: {sheet_name}")
        # Stream rows; keep readable plaintext (tabs between cells, newlines between rows).
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                chunks.append("\t".join(cells))
        chunks.append("")  # blank line between sheets

    text = "\n".join(chunks).strip()
    meta = {"sheet_names": sheet_names, "num_sheets": len(sheet_names)}
    return text, meta


def _load_json_docs(path: Path) -> List[Dict[str, Any]]:
    """
    Load a JSON file and return a list of document dicts ready for
    canonicalization.

    Pipeline:
      1. Try the repo's ``load_article_json`` (handles many common shapes).
      2. Fall back to ``_try_repair_json`` for hand-edited / malformed files.
      3. Fall back to ``_extract_json_objects_from_text`` — individually
         parse each top-level ``{…}`` block when the overall array is broken.
      4. Detect and flatten press / news-style nested documents.
    """
    docs: Optional[List[Dict[str, Any]]] = None
    raw_text = path.read_text(encoding="utf-8", errors="replace")

    # --- attempt 1: standard parser ---
    try:
        parsed = load_article_json(str(path))
        if isinstance(parsed, list) and parsed:
            docs = parsed
    except Exception:
        pass

    # --- attempt 2: repair + parse ---
    if docs is None:
        try:
            data = _try_repair_json(raw_text)
            if isinstance(data, list):
                docs = [d for d in data if isinstance(d, dict)]
            elif isinstance(data, dict):
                docs = [data]
        except Exception:
            pass

    # --- attempt 3: extract individual JSON objects from text ---
    if not docs:
        docs = _extract_json_objects_from_text(raw_text)

    if not docs:
        raise ValueError(f"No document records found in JSON file: {path}")

    # --- flatten press / news-article documents ---
    return [
        _flatten_press_doc(d, doc_index=i) if _is_press_style_doc(d) else d
        for i, d in enumerate(docs, start=1)
    ]


def _load_jsonl_docs(path: Path) -> List[Dict[str, Any]]:
    docs: List[Dict[str, Any]] = []
    with open(path, "r", encoding="utf-8") as f:
        for line_num, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError as e:
                raise json.JSONDecodeError(
                    f"Invalid JSON on line {line_num} of {path}: {e.msg}", e.doc, e.pos
                )
            if isinstance(obj, dict):
                docs.append(obj)
    if not docs:
        raise ValueError(f"No documents found in JSONL file: {path}")
    return docs


def _write_jsonl(records: Iterable[Dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        for rec in records:
            json.dump(rec, f, ensure_ascii=False)
            f.write("\n")


def _detect_input_type(input_type: Optional[str], input_path: Path) -> str:
    if input_type:
        normalized = input_type.strip().lower()
        if normalized not in SUPPORTED_INPUT_TYPES:
            raise ValueError(f"Unsupported --input-type {input_type!r}. Use one of: {SUPPORTED_INPUT_TYPES}")
        return normalized

    suffix = input_path.suffix.lower().lstrip(".")
    if suffix in SUPPORTED_INPUT_TYPES:
        return suffix
    raise ValueError(
        f"Cannot infer input type from extension: {input_path.suffix!r}. "
        f"Provide --input-type ({', '.join(SUPPORTED_INPUT_TYPES)})."
    )


def convert_to_qagredo_jsonl(input_file: str, output_file: str, input_type: Optional[str] = None) -> int:
    input_path = _resolve_input_path(input_file)
    inferred_type = _detect_input_type(input_type, input_path)

    source_str = str(input_path)
    stem = _safe_slug(input_path.stem)

    records: List[Dict[str, Any]] = []

    if inferred_type == "txt":
        text = _extract_txt(input_path)
        raw = {"id": stem, "title": input_path.stem, "content": text, "source": source_str, "type": "text_document"}
        records.append(_canonicalize_record(raw, fallback_id=stem, fallback_title=input_path.stem, source=source_str, input_type="txt"))

    elif inferred_type == "pdf":
        text = _extract_pdf(input_path)
        raw = {"id": stem, "title": input_path.stem, "content": text, "source": source_str, "type": "text_document"}
        records.append(_canonicalize_record(raw, fallback_id=stem, fallback_title=input_path.stem, source=source_str, input_type="pdf"))

    elif inferred_type == "xlsx":
        text, meta = _extract_xlsx(input_path)
        raw = {"id": stem, "title": input_path.stem, "content": text, "source": source_str, "type": "text_document"}
        records.append(
            _canonicalize_record(
                raw,
                fallback_id=stem,
                fallback_title=input_path.stem,
                source=source_str,
                input_type="xlsx",
                extra_metadata=meta,
            )
        )

    elif inferred_type == "json":
        docs = _load_json_docs(input_path)
        for idx, doc in enumerate(docs, start=1):
            doc_id = str(doc.get("id") or f"{stem}_{idx}" if len(docs) > 1 else stem)
            title = str(doc.get("title") or (f"{input_path.stem} {idx}" if len(docs) > 1 else input_path.stem))
            records.append(
                _canonicalize_record(
                    doc,
                    fallback_id=doc_id,
                    fallback_title=title,
                    source=source_str,
                    input_type="json",
                    extra_metadata={"source_path": source_str},
                )
            )

    elif inferred_type == "jsonl":
        docs = _load_jsonl_docs(input_path)
        for idx, doc in enumerate(docs, start=1):
            doc_id = str(doc.get("id") or f"{stem}_{idx}" if len(docs) > 1 else stem)
            title = str(doc.get("title") or (f"{input_path.stem} {idx}" if len(docs) > 1 else input_path.stem))
            records.append(
                _canonicalize_record(
                    doc,
                    fallback_id=doc_id,
                    fallback_title=title,
                    source=source_str,
                    input_type="jsonl",
                    extra_metadata={"source_path": source_str, "line_number": idx},
                )
            )

    else:
        raise ValueError(f"Unsupported input type: {inferred_type}")

    output_path = Path(output_file)
    _write_jsonl(records, output_path)

    total_chars = sum(len(r.get("content", "")) for r in records)
    print("=" * 80)
    print(f"✅ Script version: 2026-02-10 (content from article, not summary)")
    print(f"✅ Converted: {input_path}")
    print(f"✅ Input type: {inferred_type}")
    print(f"✅ Wrote JSONL: {output_path} ({len(records)} record(s), {total_chars} chars total)")
    for i, r in enumerate(records, 1):
        meta = r.get("metadata", {})
        src_count = meta.get("num_sources", 1)
        has_summary = "summary" in meta
        print(f"   Record {i}: title={r.get('title','?')!r}, sources={src_count}, summary_in_metadata={has_summary}")
    print("=" * 80)
    return 0


def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Convert pdf/txt/xlsx/json/jsonl into QAGRedo-compatible JSONL.")
    p.add_argument("--input", required=True, help="Input file path (pdf/txt/xlsx/json/jsonl)")
    p.add_argument("--output", required=True, help="Output JSONL path")
    p.add_argument(
        "--input-type",
        default=None,
        choices=SUPPORTED_INPUT_TYPES,
        help="Override input type (otherwise inferred from extension).",
    )
    return p


def main(argv: Optional[List[str]] = None) -> int:
    args = _build_arg_parser().parse_args(argv)
    return convert_to_qagredo_jsonl(args.input, args.output, input_type=args.input_type)


if __name__ == "__main__":
    raise SystemExit(main())

